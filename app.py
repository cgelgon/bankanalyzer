import os, io, json, re, unicodedata
from datetime import datetime
from collections import Counter
from concurrent.futures import ThreadPoolExecutor, as_completed
from flask import Flask, request, jsonify, redirect, Response
from flask_cors import CORS
import pypdf
import anthropic
import openpyxl
import stripe
import psycopg2
from psycopg2.extras import RealDictCursor

app = Flask(__name__)
CORS(app)

stripe.api_key = os.environ.get('STRIPE_SECRET_KEY', '')
STRIPE_PRICE_ID = os.environ.get('STRIPE_PRICE_ID', '')
STRIPE_WEBHOOK_SECRET = os.environ.get('STRIPE_WEBHOOK_SECRET', '')
FRONTEND_URL = os.environ.get('FRONTEND_URL', 'https://bankanalyzer-nu.vercel.app')
DATABASE_URL = os.environ.get('DATABASE_URL', '')


def get_db():
    conn = psycopg2.connect(DATABASE_URL, cursor_factory=RealDictCursor)
    return conn


def init_db():
    if not DATABASE_URL:
        print('DATABASE_URL non definie, la base ne sera pas initialisee')
        return
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute('''
            CREATE TABLE IF NOT EXISTS users (
                email TEXT PRIMARY KEY,
                stripe_customer_id TEXT,
                stripe_subscription_id TEXT,
                subscription_status TEXT DEFAULT 'inactive',
                created_at TIMESTAMP DEFAULT NOW(),
                updated_at TIMESTAMP DEFAULT NOW()
            )
        ''')
        cur.execute('ALTER TABLE users ADD COLUMN IF NOT EXISTS nb_analyses_mois_courant INTEGER DEFAULT 0')
        cur.execute('ALTER TABLE users ADD COLUMN IF NOT EXISTS mois_reference_quota TEXT')
        cur.execute('''
            CREATE TABLE IF NOT EXISTS banques_detectees (
                id SERIAL PRIMARY KEY,
                nom_banque TEXT,
                format TEXT,
                periode TEXT,
                created_at TIMESTAMP DEFAULT NOW()
            )
        ''')
        conn.commit()
        cur.close()
        conn.close()
        print('Base de donnees initialisee avec succes')
    except Exception as e:
        print('ERREUR init_db:', str(e))


def est_pro(email):
    if not email or not DATABASE_URL:
        return False
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute('SELECT subscription_status FROM users WHERE email = %s', (email.lower().strip(),))
        row = cur.fetchone()
        cur.close()
        conn.close()
        return bool(row and row['subscription_status'] == 'active')
    except Exception as e:
        print('ERREUR est_pro:', str(e))
        return False


QUOTA_ANALYSES_PRO_PAR_MOIS = 5


def verifier_et_incrementer_quota_pro(email):
    """Verifie que le compte Pro n'a pas depasse son quota d'analyses
    pour le mois en cours, et incremente son compteur si l'analyse est
    autorisee. Le quota se reinitialise automatiquement a chaque nouveau
    mois calendaire. Sert de garde-fou contre le partage d'un meme compte
    Pro entre plusieurs personnes (voir discussion du 08/08/2026).

    Retourne (autorise: bool, nb_utilisees: int).
    """
    if not email or not DATABASE_URL:
        return True, 0  # pas de DB configuree -> on n'entrave pas l'usage

    mois_actuel = datetime.now().strftime('%Y-%m')
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute(
            'SELECT nb_analyses_mois_courant, mois_reference_quota FROM users WHERE email = %s',
            (email.lower().strip(),)
        )
        row = cur.fetchone()
        if row is None:
            cur.close()
            conn.close()
            return True, 0

        nb_actuel = row['nb_analyses_mois_courant'] or 0
        mois_enregistre = row['mois_reference_quota']
        if mois_enregistre != mois_actuel:
            # nouveau mois calendaire : on reinitialise le compteur
            nb_actuel = 0

        if nb_actuel >= QUOTA_ANALYSES_PRO_PAR_MOIS:
            cur.close()
            conn.close()
            return False, nb_actuel

        nouveau_nb = nb_actuel + 1
        cur.execute(
            'UPDATE users SET nb_analyses_mois_courant = %s, mois_reference_quota = %s, updated_at = NOW() WHERE email = %s',
            (nouveau_nb, mois_actuel, email.lower().strip())
        )
        conn.commit()
        cur.close()
        conn.close()
        return True, nouveau_nb
    except Exception as e:
        print('ERREUR verifier_et_incrementer_quota_pro:', str(e))
        return True, 0  # en cas d'erreur DB, on n'entrave pas l'usage (fail-open)

def _logger_banque_detectee(nom_banque, nom_fichier, periode):
    """Enregistre discretement la banque detectee par l'IA pour chaque
    analyse, sans rien demander a l'utilisateur -- sert uniquement a
    prioriser plus tard les formats meritant un parseur dedie (comme
    celui deja fait pour Revolut), en fonction de l'usage reel observe.
    """
    if not DATABASE_URL:
        return
    extension = (nom_fichier or '').lower().rsplit('.', 1)[-1] if '.' in (nom_fichier or '') else '?'
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute(
            'INSERT INTO banques_detectees (nom_banque, format, periode) VALUES (%s, %s, %s)',
            ((nom_banque or 'Inconnu')[:100], extension, periode or '')
        )
        conn.commit()
        cur.close()
        conn.close()
    except Exception as e:
        print('ERREUR log banque detectee:', str(e))

def upsert_user(email, stripe_customer_id=None, stripe_subscription_id=None, subscription_status=None):
    if not DATABASE_URL:
        return
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute('SELECT email FROM users WHERE email = %s', (email.lower().strip(),))
        existe = cur.fetchone()
        if existe:
            champs = []
            valeurs = []
            if stripe_customer_id is not None:
                champs.append('stripe_customer_id = %s')
                valeurs.append(stripe_customer_id)
            if stripe_subscription_id is not None:
                champs.append('stripe_subscription_id = %s')
                valeurs.append(stripe_subscription_id)
            if subscription_status is not None:
                champs.append('subscription_status = %s')
                valeurs.append(subscription_status)
            champs.append('updated_at = NOW()')
            valeurs.append(email.lower().strip())
            cur.execute('UPDATE users SET ' + ', '.join(champs) + ' WHERE email = %s', tuple(valeurs))
        else:
            cur.execute(
                'INSERT INTO users (email, stripe_customer_id, stripe_subscription_id, subscription_status) VALUES (%s, %s, %s, %s)',
                (email.lower().strip(), stripe_customer_id, stripe_subscription_id, subscription_status or 'inactive')
            )
        conn.commit()
        cur.close()
        conn.close()
    except Exception as e:
        print('ERREUR upsert_user:', str(e))


init_db()

MOIS_MAP = {
    # francais
    'janvier': 1, 'fevrier': 2, 'mars': 3, 'avril': 4, 'mai': 5, 'juin': 6,
    'juillet': 7, 'aout': 8, 'septembre': 9, 'octobre': 10, 'novembre': 11, 'decembre': 12,
    # english
    'january': 1, 'february': 2, 'march': 3, 'april': 4, 'may': 5, 'june': 6,
    'july': 7, 'august': 8, 'september': 9, 'october': 10, 'november': 11, 'december': 12,
    # espanol
    'enero': 1, 'febrero': 2, 'marzo': 3, 'abril': 4, 'mayo': 5, 'junio': 6,
    'julio': 7, 'agosto': 8, 'septiembre': 9, 'setiembre': 9, 'octubre': 10, 'noviembre': 11, 'diciembre': 12,
    # deutsch
    'januar': 1, 'februar': 2, 'marz': 3, 'juni': 6, 'juli': 7,
    'oktober': 10, 'dezember': 12,
    # italiano
    'gennaio': 1, 'febbraio': 2, 'aprile': 4, 'maggio': 5, 'giugno': 6, 'luglio': 7,
    'settembre': 9, 'ottobre': 10, 'dicembre': 12,
    # portugues
    'janeiro': 1, 'fevereiro': 2, 'marco': 3, 'maio': 5, 'junho': 6, 'julho': 7,
    'setembro': 9, 'outubro': 10, 'novembro': 11, 'dezembro': 12,
    # arabe
    'يناير': 1, 'فبراير': 2, 'مارس': 3, 'أبريل': 4, 'مايو': 5, 'يونيو': 6,
    'يوليو': 7, 'أغسطس': 8, 'سبتمبر': 9, 'أكتوبر': 10, 'نوفمبر': 11, 'ديسمبر': 12,
}


def sans_accents(s):
    return ''.join(c for c in unicodedata.normalize('NFD', s) if unicodedata.category(c) != 'Mn')


def to_num(v):
    """Convertit une valeur (potentiellement une chaine renvoyee par l'IA) en nombre, sans jamais planter."""
    if isinstance(v, (int, float)):
        return v
    if v is None:
        return 0.0
    try:
        s = str(v).strip().replace('\u202f', '').replace(' ', '').replace(',', '.')
        return float(s)
    except (TypeError, ValueError):
        return 0.0


_REGEX_5_MONTANTS_EUR = re.compile(
    r'(-?\d[\d\s\u202f]*,\d{2})€\s*(-?\d[\d\s\u202f]*,\d{2})€\s*(-?\d[\d\s\u202f]*,\d{2})€\s*(-?\d[\d\s\u202f]*,\d{2})€\s*(-?\d[\d\s\u202f]*,\d{2})€'
)


def _extraire_soldes_ouverture_cloture_texte(text):
    """Cherche 'Solde d'ouverture' / 'Solde de cloture' dans un texte
    de releve (present et fiable sur le format Revolut)."""
    m_ouv = re.search(r"Solde d'ouverture\s*(-?\d[\d\s\u202f]*,\d{2})\s*€", text)
    m_clo = re.search(r"Solde de cl[oô]ture\s*(-?\d[\d\s\u202f]*,\d{2})\s*€", text)
    if m_ouv and m_clo:
        return to_num(m_ouv.group(1)), to_num(m_clo.group(1))
    return None, None


def calculer_totaux_verifies_pdf(text):
    """Essaie de calculer des totaux exacts pour certains formats de PDF
    reconnus (actuellement : Revolut 'Releve personnalise'), plutot que
    de laisser l'IA les estimer. Retourne None si le format n'est pas
    reconnu (comportement inchange, aucune regression)."""
    if 'Revolut' in text and ('Relevé des transactions' in text or ('Argent' in text and 'entrant' in text)):
        matches = _REGEX_5_MONTANTS_EUR.findall(text)
        if matches:
            recettes = 0.0
            depenses = 0.0
            for groupe in matches:
                m = to_num(groupe[0])
                if m >= 0:
                    recettes += m
                else:
                    depenses += -m
            totaux = {'totalRecettes': round(recettes, 2), 'totalDepenses': round(depenses, 2)}
            solde_ouv, solde_clo = _extraire_soldes_ouverture_cloture_texte(text)
            if solde_ouv is not None:
                totaux['soldeDepart'] = solde_ouv
                totaux['soldeArrivee'] = solde_clo
            return totaux
    return None

def extract_text_from_pdf(file_bytes):
    text = ''
    reader = pypdf.PdfReader(io.BytesIO(file_bytes))
    for page in reader.pages:
        t = page.extract_text()
        if t:
            text += t + chr(10)
    return text


def extract_text_from_csv(file_bytes):
    for encodage in ('utf-8-sig', 'utf-8', 'cp1252', 'latin-1'):
        try:
            return file_bytes.decode(encodage)
        except (UnicodeDecodeError, LookupError):
            continue
    return file_bytes.decode('utf-8', errors='ignore')


def extract_text_from_excel(file_bytes):
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
        ws = wb.active
        lines = []
        for row in ws.iter_rows(values_only=True):
            lines.append(' | '.join([str(c) if c else '' for c in row]))
        return chr(10).join(lines)
    except Exception:
        # Ancien format .xls (binaire, pas gere par openpyxl) : on tente avec xlrd
        import xlrd
        wb = xlrd.open_workbook(file_contents=file_bytes)
        ws = wb.sheet_by_index(0)
        lines = []
        for r in range(ws.nrows):
            row = ws.row_values(r)
            lines.append(' | '.join([str(c) if c not in (None, '') else '' for c in row]))
        return chr(10).join(lines)


def _trouver_index_colonne_date(entetes):
    """Cherche l'index de la colonne de date la plus fiable (Date de
    comptabilisation en priorite), en tolerant variations d'accents/casse."""
    priorite = [
        'date de comptabilisation', 'date operation', "date d'operation",
        'date valeur', 'date de valeur', 'date transaction', 'date',
    ]
    entetes_norm = [sans_accents(str(h or '').strip().lower()) for h in entetes]
    for candidat in priorite:
        candidat_norm = sans_accents(candidat)
        for i, e in enumerate(entetes_norm):
            if candidat_norm == e or candidat_norm in e:
                return i
    return None


def _trouver_index_colonne_contrepartie(entetes):
    """Cherche l'index de la colonne identifiant le tiers d'une transaction
    (contrepartie/beneficiaire/libelle), en tolerant variations
    d'accents/casse. Retourne None si aucune colonne fiable n'est trouvee --
    dans ce cas l'appelant DOIT retomber sur le comportement actuel, jamais
    deviner."""
    priorite = [
        'contrepartie', 'beneficiaire', 'tiers', 'nom du tiers',
        'libelle operation', "libelle de l'operation", 'libelle',
        'description', 'intitule',
    ]
    entetes_norm = [sans_accents(str(h or '').strip().lower()) for h in entetes]
    for candidat in priorite:
        candidat_norm = sans_accents(candidat)
        for i, e in enumerate(entetes_norm):
            if candidat_norm == e or candidat_norm in e:
                return i
    return None


def _parser_date_cellule(valeur):
    """Convertit une valeur de cellule (datetime deja parse par openpyxl,
    ou chaine issue d'un CSV) en tuple (annee, mois), ou None si non reconnue."""
    if isinstance(valeur, datetime):
        return (valeur.year, valeur.month)
    if isinstance(valeur, str):
        v = valeur.strip()
        for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%d.%m.%Y', '%Y/%m/%d'):
            try:
                d = datetime.strptime(v[:10], fmt)
                return (d.year, d.month)
            except ValueError:
                continue
    return None


def _trouver_colonnes_montant(entetes):
    """Cherche les colonnes Debit/Credit (ou a defaut une colonne Montant
    signee) pour pouvoir calculer les totaux exacts en Python plutot que
    de laisser l'IA les estimer."""
    entetes_norm = [sans_accents(str(h or '').strip().lower()) for h in entetes]
    idx_debit = None
    idx_credit = None
    idx_montant = None
    for i, e in enumerate(entetes_norm):
        if idx_debit is None and ('debit' in e):
            idx_debit = i
        elif idx_credit is None and ('credit' in e):
            idx_credit = i
        elif idx_montant is None and e == 'montant':
            idx_montant = i
    return idx_debit, idx_credit, idx_montant


def _totaux_ligne(ligne, idx_debit, idx_credit, idx_montant):
    """Retourne (recette, depense) pour une ligne de donnees, calcules
    a partir des colonnes Debit/Credit ou Montant identifiees."""
    recette = 0.0
    depense = 0.0
    if idx_credit is not None and idx_credit < len(ligne):
        recette = to_num(ligne[idx_credit])
    if idx_debit is not None and idx_debit < len(ligne):
        depense = abs(to_num(ligne[idx_debit]))
    if idx_credit is None and idx_debit is None and idx_montant is not None and idx_montant < len(ligne):
        m = to_num(ligne[idx_montant])
        if m >= 0:
            recette = m
        else:
            depense = abs(m)
    return recette, depense


def _decouper_lignes_par_mois(entetes, lignes_donnees):
    """Regroupe des lignes de donnees (CSV ou Excel) par mois reellement
    detecte dans la colonne de date la plus fiable, ET calcule en Python
    les totaux recettes/depenses exacts de chaque mois (voir
    _trouver_colonnes_montant). Retourne une liste de blocs
    {'periode': 'MM/AAAA' ou None, 'text': texte_du_bloc,
    'totauxVerifies': {'totalRecettes': x, 'totalDepenses': y} ou None}.

    Si aucune colonne de date n'est identifiable, retourne UN SEUL bloc.
    Si aucune colonne Debit/Credit/Montant n'est identifiable,
    'totauxVerifies' vaut None (comportement inchange, l'IA estime comme avant).
    """
    ligne_entete_txt = ' | '.join([str(c) if c not in (None, '') else '' for c in entetes])
    idx_date = _trouver_index_colonne_date(entetes)
    idx_debit, idx_credit, idx_montant = _trouver_colonnes_montant(entetes)
    colonnes_montant_ok = idx_debit is not None or idx_credit is not None or idx_montant is not None
    idx_cp = _trouver_index_colonne_contrepartie(entetes)

    # On retire les transactions annulees AVANT tout regroupement : elles ne
    # se sont jamais reellement produites, ne doivent pas compter dans les
    # totaux, et ne doivent plus former un bloc "periode inconnue" a part.
    idx_etat = None
    for i, entete in enumerate(entetes):
        if sans_accents(str(entete or '').strip().lower()) == 'etat':
            idx_etat = i
            break
    if idx_etat is not None:
        lignes_donnees = [
            ligne for ligne in lignes_donnees
            if not (idx_etat < len(ligne) and sans_accents(str(ligne[idx_etat] or '').strip().lower()) == 'annule')
        ]

    def _extraire_tx_structuree(ligne):
        # Retourne un dict {'contrepartie','recette','depense'} exploitable
        # en Python (sans passer par l'IA), ou None si pas exploitable pour
        # cette ligne -- jamais de valeur inventee.
        if idx_cp is None or not colonnes_montant_ok:
            return None
        contrepartie = str(ligne[idx_cp]).strip() if idx_cp < len(ligne) and ligne[idx_cp] else ''
        if not contrepartie:
            return None
        r, d = _totaux_ligne(ligne, idx_debit, idx_credit, idx_montant)
        if r == 0 and d == 0:
            return None
        return {'contrepartie': contrepartie, 'recette': round(r, 2), 'depense': round(d, 2)}

    if idx_date is None:
        lignes_txt = [ligne_entete_txt] + [
            ' | '.join([str(c) if c not in (None, '') else '' for c in ligne]) for ligne in lignes_donnees
        ]
        totaux = None
        if colonnes_montant_ok:
            r_tot, d_tot = 0.0, 0.0
            for ligne in lignes_donnees:
                r, d = _totaux_ligne(ligne, idx_debit, idx_credit, idx_montant)
                r_tot += r
                d_tot += d
            totaux = {'totalRecettes': round(r_tot, 2), 'totalDepenses': round(d_tot, 2)}
        transactions_structurees = None
        if idx_cp is not None and colonnes_montant_ok:
            transactions_structurees = [t for t in (_extraire_tx_structuree(l) for l in lignes_donnees) if t]
        return [{'periode': None, 'text': chr(10).join(lignes_txt), 'totauxVerifies': totaux, 'transactionsStructurees': transactions_structurees}]

    groupes = {}
    totaux_groupes = {}
    tx_structurees_groupes = {}
    ordre_apparition = []
    for ligne in lignes_donnees:
        valeur_date = ligne[idx_date] if idx_date < len(ligne) else None
        cle = _parser_date_cellule(valeur_date) or ('inconnue', 0)
        if cle not in groupes:
            groupes[cle] = []
            totaux_groupes[cle] = [0.0, 0.0]
            tx_structurees_groupes[cle] = []
            ordre_apparition.append(cle)
        groupes[cle].append(' | '.join([str(c) if c not in (None, '') else '' for c in ligne]))
        if colonnes_montant_ok:
            r, d = _totaux_ligne(ligne, idx_debit, idx_credit, idx_montant)
            totaux_groupes[cle][0] += r
            totaux_groupes[cle][1] += d
        tx = _extraire_tx_structuree(ligne)
        if tx:
            tx_structurees_groupes[cle].append(tx)

    blocs = []
    for cle in sorted(ordre_apparition, key=lambda c: (str(c[0]), c[1])):
        annee, mois = cle
        periode = ('%02d/%s' % (mois, annee)) if mois else None
        texte_bloc = chr(10).join([ligne_entete_txt] + groupes[cle])
        totaux = None
        if colonnes_montant_ok:
            r_tot, d_tot = totaux_groupes[cle]
            totaux = {'totalRecettes': round(r_tot, 2), 'totalDepenses': round(d_tot, 2)}
        transactions_structurees = tx_structurees_groupes[cle] if (idx_cp is not None and colonnes_montant_ok) else None
        blocs.append({'periode': periode, 'text': texte_bloc, 'totauxVerifies': totaux, 'transactionsStructurees': transactions_structurees})

    return blocs


def decouper_par_mois_excel(file_bytes):
    """Version 'decoupage par mois' de extract_text_from_excel : au lieu
    d'un seul bloc de texte pour tout le classeur, retourne une liste de
    blocs, un par mois reellement detecte (voir _decouper_lignes_par_mois).
    """
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
        ws = wb.active
        toutes_lignes = list(ws.iter_rows(values_only=True))
    except Exception:
        import xlrd
        wb = xlrd.open_workbook(file_contents=file_bytes)
        ws = wb.sheet_by_index(0)
        toutes_lignes = [tuple(ws.row_values(r)) for r in range(ws.nrows)]

    if not toutes_lignes:
        return [{'periode': None, 'text': ''}]

    return _decouper_lignes_par_mois(toutes_lignes[0], toutes_lignes[1:])


def _detecter_delimiteur_csv(texte_brut):
    """Comme csv.Sniffer(), mais avec un filet de securite : si le Sniffer
    echoue (frequent sur les exports bancaires reels), on teste nous-memes
    les delimiteurs candidats en comptant leurs occurrences sur les
    premieres lignes, et on retient celui dont le compte est non nul ET
    constant sur toutes les lignes -- signe fiable du vrai separateur.
    Sans ca, le comportement par defaut de Python bascule sur la virgule,
    ce qui casse silencieusement tout export francais (point-virgule).
    """
    import csv as _csv
    try:
        return _csv.Sniffer().sniff(texte_brut[:4096], delimiters=';,\t')
    except Exception:
        pass
    lignes_echantillon = [l for l in texte_brut.splitlines()[:5] if l.strip()]
    meilleur_delim = ';'
    meilleur_score = 0
    for delim in (';', ',', '\t'):
        comptes = [ligne.count(delim) for ligne in lignes_echantillon]
        if comptes and min(comptes) > 0 and len(set(comptes)) == 1:
            if comptes[0] > meilleur_score:
                meilleur_score = comptes[0]
                meilleur_delim = delim

    class _DialecteManuel(_csv.excel):
        delimiter = meilleur_delim

    return _DialecteManuel


def decouper_par_mois_csv(file_bytes):
    """Version 'decoupage par mois' pour les CSV : parse les lignes (au
    lieu de renvoyer le texte brut tel quel) et regroupe par mois reellement
    detecte, comme decouper_par_mois_excel.
    """
    import csv as _csv
    texte_brut = extract_text_from_csv(file_bytes)
    dialecte = _detecter_delimiteur_csv(texte_brut)
    lecteur = _csv.reader(io.StringIO(texte_brut), dialecte)
    toutes_lignes = [tuple(ligne) for ligne in lecteur if ligne]

    if not toutes_lignes:
        return [{'periode': None, 'text': ''}]

    return _decouper_lignes_par_mois(toutes_lignes[0], toutes_lignes[1:])

_REGEX_DATE_ISO_DATETIME = re.compile(r'\b(\d{4})-(\d{2})-(\d{2}) \d{2}:\d{2}:\d{2}\b')


def get_periode(text):
    # PRIORITE 1 : dates ISO issues des vraies colonnes de date (CSV/XLSX).
    # str(datetime) produit "YYYY-MM-DD HH:MM:SS" -- format fiable qui
    # n'apparait jamais par hasard dans du texte libre (notes, references
    # de facture type "Juin 2026", "TSP 06.2026", etc.). On prend le mois
    # le PLUS FREQUENT parmi toutes les dates trouvees, jamais la 1ere.
    matches_iso = _REGEX_DATE_ISO_DATETIME.findall(text)
    if matches_iso:
        mois_annees = [(int(a), int(m)) for a, m, j in matches_iso if 1 <= int(m) <= 12]
        if mois_annees:
            (annee, mois), _occurrences = Counter(mois_annees).most_common(1)[0]
            return f"{mois:02d}/{annee}"

    # ---- reste du code EXISTANT inchange (fallback pour PDF texte pur) ----
    # Format chinois : "2026年7月" ou "7月2026年"
    m_cn = re.search(r'(\d{4})\s*年\s*(\d{1,2})\s*月', text)
    if m_cn:
        return m_cn.group(2) + '/' + m_cn.group(1)
    m_cn2 = re.search(r'(\d{1,2})\s*月\s*(\d{4})\s*年', text)
    if m_cn2:
        return m_cn2.group(1) + '/' + m_cn2.group(2)

    texte_sans_accents = sans_accents(text.lower())
    noms_mois = '|'.join(sorted(MOIS_MAP.keys(), key=len, reverse=True))
    m = re.search('(' + noms_mois + r')\s*\d{4}', texte_sans_accents)
    if m:
        return m.group(0).capitalize()

    m2 = re.search(r'(\d{1,2})[/-](\d{4})', text)
    if m2:
        return m2.group(0)
    return 'Periode inconnue'


def periode_sort_key(periode):
    """Retourne une cle (annee, mois) triable a partir d'une chaine periode. (0,0) si non reconnue."""
    if not periode:
        return (0, 0)
    p_low = sans_accents(periode.lower())
    for nom, num in MOIS_MAP.items():
        m = re.search(nom + r'\s*(\d{4})', p_low)
        if m:
            return (int(m.group(1)), num)
    m2 = re.search(r'(\d{1,2})[/-](\d{4})', periode)
    if m2:
        return (int(m2.group(2)), int(m2.group(1)))
    return (0, 0)


LIMITE_CARACTERES_RELEVE = 500000  # marge large ; le decoupage par mois (bug 2)
# limite deja la taille par bloc, ceci couvre les mois a fort volume (500+ transactions)


def _verifier_taille_et_tronquer(text):
    """Remplace l'ancienne troncature silencieuse text[:20000].
    Si le texte tient dans la nouvelle limite (large), on l'envoie tel quel.
    S'il la depasse quand meme (fichier tres volumineux), on leve une
    erreur EXPLICITE plutot que d'analyser silencieusement une fraction
    des donnees et de presenter un resultat faux avec un score de confiance.
    """
    if len(text) <= LIMITE_CARACTERES_RELEVE:
        return text
    print('ATTENTION: fichier trop volumineux pour une analyse complete -- '
          + str(len(text)) + ' caracteres, limite ' + str(LIMITE_CARACTERES_RELEVE))
    raise ValueError(
        'Ce fichier est trop volumineux pour etre analyse en une seule fois '
        '(' + str(len(text)) + ' caracteres, limite actuelle ' + str(LIMITE_CARACTERES_RELEVE) + '). '
        'Merci de le scinder en plusieurs fichiers (par mois ou par compte) avant de le renvoyer.'
    )


SEUIL_TRANSACTIONS_CATEGORISATION_EXHAUSTIVE = 300


def categoriser_contreparties_par_categories(client, langue, categories_recettes, categories_depenses, contreparties_recettes, contreparties_depenses):
    """Appel IA dedie et peu couteux : assigne chaque contrepartie unique a
    UNE des categories DEJA DECIDEES par l'analyse principale (jamais de
    nouvelle categorie inventee -- garantit la coherence avec ce qui est
    deja affiche a l'ecran). Beaucoup moins couteux qu'une categorisation
    transaction par transaction sur les gros relevees."""
    if not contreparties_recettes and not contreparties_depenses:
        return {'mapping_recettes': {}, 'mapping_depenses': {}}
    prompt = (
        'INSTRUCTION ABSOLUE: reponds UNIQUEMENT en JSON valide, sans markdown, sans texte avant/apres.' + chr(10) +
        'Categories de RECETTES deja identifiees (n\'en invente JAMAIS d\'autres, recopie a l\'identique) : ' + json.dumps(categories_recettes, ensure_ascii=False) + chr(10) +
        'Categories de DEPENSES deja identifiees (n\'en invente JAMAIS d\'autres, recopie a l\'identique) : ' + json.dumps(categories_depenses, ensure_ascii=False) + chr(10) +
        'Contreparties RECETTES a classer : ' + json.dumps(contreparties_recettes, ensure_ascii=False) + chr(10) +
        'Contreparties DEPENSES a classer : ' + json.dumps(contreparties_depenses, ensure_ascii=False) + chr(10) +
        'Pour CHAQUE contrepartie listee, assigne EXACTEMENT une categorie parmi celles fournies ci-dessus (recopie le libelle a l\'identique, caractere pour caractere). ' +
        'Si aucune ne convient clairement, utilise "Autres". Reponds UNIQUEMENT avec ce JSON :' + chr(10) +
        '{"mapping_recettes":{"nom_contrepartie":"categorie"},"mapping_depenses":{"nom_contrepartie":"categorie"}}'
    )
    msg = client.messages.create(
        model='claude-sonnet-4-6',
        max_tokens=16000,
        system='Tu classes des contreparties bancaires dans des categories deja fournies. Tu ne reponds qu avec du JSON valide, rien d autre, sans markdown.',
        messages=[{'role': 'user', 'content': prompt}],
    )
    texte = ''.join(b.text for b in msg.content if hasattr(b, 'text')).strip()
    if texte.startswith('```'):
        texte = texte.split(chr(10), 1)[1] if chr(10) in texte else texte
        texte = texte.rsplit('```', 1)[0]
    return json.loads(texte)


def _appliquer_categorisation_exhaustive_si_necessaire(client, f, data, langue):
    """Si ce bloc a suffisamment de transactions structurees (CSV/XLSX avec
    colonne contrepartie detectee -- voir _decouper_lignes_par_mois),
    remplace les listes de transactions tronquees (5 par categorie, voir
    analyse_releve) par des listes EXHAUSTIVES.

    SECURITE : best-effort total. Toute erreur laisse `data` EXACTEMENT
    inchange -- ne casse jamais l'analyse principale. En dessous du seuil,
    ne fait strictement rien (comportement actuel preserve)."""
    transactions_structurees = f.get('transactionsStructurees')
    if not transactions_structurees or len(transactions_structurees) < SEUIL_TRANSACTIONS_CATEGORISATION_EXHAUSTIVE:
        return
    try:
        categories_recettes = [c.get('label') for c in (data.get('recettes') or []) if c.get('label')]
        categories_depenses = [c.get('label') for c in (data.get('depenses') or []) if c.get('label')]
        if not categories_recettes and not categories_depenses:
            return

        contreparties_recettes = sorted({t['contrepartie'] for t in transactions_structurees if t.get('recette', 0) > 0})
        contreparties_depenses = sorted({t['contrepartie'] for t in transactions_structurees if t.get('depense', 0) > 0})

        mapping = categoriser_contreparties_par_categories(
            client, langue, categories_recettes, categories_depenses,
            contreparties_recettes, contreparties_depenses
        )
        mapping_recettes = mapping.get('mapping_recettes') or {}
        mapping_depenses = mapping.get('mapping_depenses') or {}

        # --- DIAGNOSTIC TEMPORAIRE ---
        manquantes_r = [c for c in contreparties_recettes if c not in mapping_recettes]
        manquantes_d = [c for c in contreparties_depenses if c not in mapping_depenses]
        autres_r = [c for c in contreparties_recettes if mapping_recettes.get(c) not in categories_recettes]
        autres_d = [c for c in contreparties_depenses if mapping_depenses.get(c) not in categories_depenses]
        print(
            'DIAGNOSTIC categorisation exhaustive -- '
            'recettes: ' + str(len(contreparties_recettes)) + ' contreparties, '
            + str(len(manquantes_r)) + ' absentes du mapping IA, '
            + str(len(autres_r)) + ' classees Autres | '
            'depenses: ' + str(len(contreparties_depenses)) + ' contreparties, '
            + str(len(manquantes_d)) + ' absentes du mapping IA, '
            + str(len(autres_d)) + ' classees Autres'
        )
        if manquantes_r:
            print('  Exemples manquantes (recettes):', manquantes_r[:5])
        if manquantes_d:
            print('  Exemples manquantes (depenses):', manquantes_d[:5])
        # --- FIN DIAGNOSTIC TEMPORAIRE ---

        recettes_par_cat = {}
        depenses_par_cat = {}
        for t in transactions_structurees:
            cp = t.get('contrepartie')
            r = t.get('recette', 0) or 0
            d = t.get('depense', 0) or 0
            if r > 0:
                cat = mapping_recettes.get(cp)
                cat = cat if cat in categories_recettes else 'Autres'
                recettes_par_cat.setdefault(cat, {'montant': 0.0, 'transactions': []})
                recettes_par_cat[cat]['montant'] += r
                recettes_par_cat[cat]['transactions'].append({'libelle': cp, 'montant': round(r, 2)})
            if d > 0:
                cat = mapping_depenses.get(cp)
                cat = cat if cat in categories_depenses else 'Autres'
                depenses_par_cat.setdefault(cat, {'montant': 0.0, 'transactions': []})
                depenses_par_cat[cat]['montant'] += d
                depenses_par_cat[cat]['transactions'].append({'libelle': cp, 'montant': round(d, 2)})

        # On ne remplace QUE les categories deja presentes dans data (jamais
        # de nouvelle categorie ajoutee a l'ecran a ce stade). Une categorie
        # sans aucune contrepartie mappee (ex. "Autres" si personne n'y a
        # ete assigne) garde son contenu original -- comportement de repli
        # volontaire, jamais de donnee effacee.
        for categorie_liste, par_cat in ((data.get('recettes'), recettes_par_cat), (data.get('depenses'), depenses_par_cat)):
            if categorie_liste is None:
                continue
            for entry in categorie_liste:
                label = entry.get('label')
                if label in par_cat:
                    entry['transactions'] = par_cat[label]['transactions']
                    entry['montant'] = round(par_cat[label]['montant'], 2)
            # Si des contreparties ont ete classees "Autres" mais qu'aucune
            # entree 'Autres' n'existe deja pour ce mois (l'IA avait rempli
            # toutes ses categories sans laisser de reste), on l'ajoute
            # plutot que de perdre ces transactions silencieusement.
            labels_existants = {e.get('label') for e in categorie_liste}
            if 'Autres' in par_cat and 'Autres' not in labels_existants:
                categorie_liste.append({
                    'label': 'Autres',
                    'montant': round(par_cat['Autres']['montant'], 2),
                    'transactions': par_cat['Autres']['transactions'],
                })
    except Exception as e:
        print('AVERTISSEMENT categorisation exhaustive ignoree (repli sur comportement actuel):', str(e))
        return


def harmoniser_categories_multimois(client, langue, labels_recettes, labels_depenses):
    """Appel IA dedie : regroupe des libelles de categories issus de
    plusieurs mois qui designent probablement le meme concept sous un nom
    canonique unique. Beaucoup moins couteux qu'une categorisation
    complete (juste des libelles courts en entree/sortie)."""
    if not labels_recettes and not labels_depenses:
        return {'mapping_recettes': {}, 'mapping_depenses': {}}
    prompt = (
        'INSTRUCTION ABSOLUE: reponds UNIQUEMENT en JSON valide, sans markdown, sans texte avant/apres.' + chr(10) +
        'Voici des libelles de categories de RECETTES issus de plusieurs mois d\'un meme releve bancaire, '
        'certains designant probablement le meme concept avec une formulation differente : '
        + json.dumps(labels_recettes, ensure_ascii=False) + chr(10) +
        'Voici des libelles de categories de DEPENSES, meme situation : '
        + json.dumps(labels_depenses, ensure_ascii=False) + chr(10) +
        'Regroupe les libelles qui designent clairement le meme concept sous UN SEUL nom canonique clair et court, '
        'vise au maximum environ 5 categories de recettes distinctes et 7 de depenses distinctes au total, sans forcer '
        'un regroupement si deux libelles designent vraiment des choses differentes. '
        'Reponds avec un mapping de CHAQUE libelle fourni (meme s\'il reste inchange) vers son nom canonique :' + chr(10) +
        '{"mapping_recettes":{"libelle_original":"nom_canonique"},"mapping_depenses":{"libelle_original":"nom_canonique"}}'
    )
    msg = client.messages.create(
        model='claude-sonnet-4-6',
        max_tokens=4000,
        system='Tu harmonises des noms de categories financieres. Tu ne reponds qu avec du JSON valide, rien d autre, sans markdown.',
        messages=[{'role': 'user', 'content': prompt}],
    )
    texte = ''.join(b.text for b in msg.content if hasattr(b, 'text')).strip()
    if texte.startswith('```'):
        texte = texte.split(chr(10), 1)[1] if chr(10) in texte else texte
        texte = texte.rsplit('```', 1)[0]
    return json.loads(texte)


def _appliquer_harmonisation_categories_si_necessaire(client, comptes, langue):
    """Si plusieurs mois produisent des noms de categories differents pour
    un meme concept, la limite de categories affichees (5 recettes/7
    depenses) peut fusionner de vraies grosses categories dans "Autres" a
    tort. Cette fonction harmonise les libelles AVANT la fusion multi-mois,
    uniquement quand c'est necessaire (plusieurs mois ET plus de labels
    distincts que la limite).

    SECURITE : best-effort total, toute erreur laisse les labels
    inchanges -- ne casse jamais l'analyse principale."""
    if len(comptes) < 2:
        return
    labels_recettes = []
    labels_depenses = []
    for c in comptes:
        for r in (c.get('recettes') or []):
            if r.get('label'):
                labels_recettes.append(r['label'])
        for d in (c.get('depenses') or []):
            if d.get('label'):
                labels_depenses.append(d['label'])
    distincts_recettes = set(labels_recettes)
    distincts_depenses = set(labels_depenses)
    if len(distincts_recettes) <= 5 and len(distincts_depenses) <= 7:
        return
    try:
        mapping = harmoniser_categories_multimois(
            client, langue, sorted(distincts_recettes), sorted(distincts_depenses)
        )
        mapping_recettes = mapping.get('mapping_recettes') or {}
        mapping_depenses = mapping.get('mapping_depenses') or {}
        for c in comptes:
            for r in (c.get('recettes') or []):
                label = r.get('label')
                if label in mapping_recettes:
                    r['label'] = mapping_recettes[label]
            for d in (c.get('depenses') or []):
                label = d.get('label')
                if label in mapping_depenses:
                    d['label'] = mapping_depenses[label]
    except Exception as e:
        print('AVERTISSEMENT harmonisation categories ignoree (repli sur labels originaux):', repr(e))
        return


def _completer_avec_categorie_autres(data):
    """Fait en sorte que la somme des categories affichees (recettes/
    depenses) corresponde toujours exactement au total affiche, SANS
    jamais depasser le nombre maximal de categories (5 recettes / 7
    depenses) pour ne pas risquer qu'une categorie soit coupee par le
    frontend :
    - Si l'IA a deja mis sa propre categorie 'Autres' (approximative),
      on remplace son montant par le reste exact.
    - Sinon, si on est sous la limite, on ajoute une categorie 'Autres'.
    - Sinon (deja au maximum), on fusionne l'ecart dans la plus petite
      categorie existante, renommee 'Autres'.
    """
    limites = {'recettes': 5, 'depenses': 7}
    for cle_total, cle_categories in (('totalRecettes', 'recettes'), ('totalDepenses', 'depenses')):
        total = to_num(data.get(cle_total, 0))
        categories = list(data.get(cle_categories) or [])
        if not categories:
            continue

        def _est_autres(c):
            return sans_accents(str(c.get('label') or '').strip().lower()) in ('autres', 'autre', 'divers')

        indices_autres = [i for i, c in enumerate(categories) if _est_autres(c)]
        if indices_autres:
            idx = indices_autres[0]
            somme_hors_autres = sum(to_num(c.get('montant', 0)) for i, c in enumerate(categories) if i != idx)
            categories = [c for i, c in enumerate(categories) if i == idx or i not in indices_autres]
            for c in categories:
                if _est_autres(c):
                    c['montant'] = round(total - somme_hors_autres, 2)
                    break
            data[cle_categories] = categories
            continue

        somme_categories = sum(to_num(c.get('montant', 0)) for c in categories)
        ecart = round(total - somme_categories, 2)
        if ecart <= 1:
            continue

        limite = limites.get(cle_categories, 999)
        if len(categories) < limite:
            categories.append({'label': 'Autres', 'montant': ecart, 'transactions': []})
        else:
            idx_min = min(range(len(categories)), key=lambda i: to_num(categories[i].get('montant', 0)))
            categories[idx_min] = {
                'label': 'Autres',
                'montant': round(to_num(categories[idx_min].get('montant', 0)) + ecart, 2),
                'transactions': [],
            }
        data[cle_categories] = categories
    return data

def analyse_releve(client, text, nom_banque, langue='français', totaux_verifies=None):
    prefixe_totaux_verifies = ''
    if totaux_verifies:
        prefixe_totaux_verifies = (
            'IMPORTANT: Les totaux suivants ont deja ete calcules avec precision a partir des '
            'donnees structurees de ce releve et DOIVENT etre utilises TELS QUELS (ne les '
            'recalcule surtout pas toi-meme) pour totalRecettes, totalDepenses, '
            'totalRecettesOfficiel et totalDepensesOfficiel, et pour toute mention de ces '
            'montants dans le commentaire, le score_detail et les actions prioritaires : '
            'Recettes = ' + str(totaux_verifies.get('totalRecettes')) + ' EUR, Depenses = ' +
            str(totaux_verifies.get('totalDepenses')) + ' EUR.' + chr(10) + chr(10)
        )
    prompt = (
        prefixe_totaux_verifies +
        'INSTRUCTION ABSOLUE: Tu dois repondre UNIQUEMENT en ' + langue + ', y compris le score_detail, le commentaire, les titres et details des actions. Aucun mot en francais si la langue demandee est differente. Analyse ce releve bancaire (' + nom_banque + ').' + chr(10) +
        'Retourne UNIQUEMENT ce JSON sans markdown:' + chr(10) +
        '{"compte":"nom de la banque et/ou du compte tel qu\'il apparait EXPLICITEMENT sur le releve (ex: REVOLUT, BNP PARIBAS - Compte Principal, Compte Booster)",' +
        '"devise":"code devise ISO du releve tel qu\'indique dessus (EUR, JOD, USD, GBP, etc.)",' +
        '"totalRecettes":0,"totalDepenses":0,"soldeDepart":0,"soldeArrivee":0,' +
        '"totalRecettesOfficiel":0,"totalDepensesOfficiel":0,' +
        '"recettes":[{"label":"cat","montant":0,"transactions":[{"libelle":"desc","montant":0,"date":"JJ/MM"}]}],' +
        '"depenses":[{"label":"cat","montant":0,"transactions":[{"libelle":"desc","montant":0,"date":"JJ/MM"}]}],' +
        '"top5depenses":[{"libelle":"desc","montant":0,"date":"JJ/MM"}],' +
        '"prelevementsRecurrents":[{"libelle":"desc","montant":0}],' +
        '"score":7,"score_detail":"phrase"}' + chr(10) +
        'REGLES:' + chr(10) +
        '1. Cherche EN PREMIER les totaux recapitulatifs (Total operations entrantes/sortantes, Solde initial, Solde final)' + chr(10) +
        '2. Utilise ces totaux pour totalRecettes et totalDepenses' + chr(10) +
        '3. soldeDepart = solde debut du releve, soldeArrivee = solde fin' + chr(10) +
        '4. top5depenses = 5 plus grosses transactions sortantes individuelles avec libelle et date' + chr(10) +
        '5. Montants entiers positifs, max 5 recettes, max 7 depenses' + chr(10) +
        '6b. Si le releve affiche un recapitulatif officiel imprime des totaux (ex: "Total des operations", "TOTAL", "Amount of Transactions", "Number/Amount of Transactions Debit/Credit"), rapporte ces totaux exacts dans totalRecettesOfficiel et totalDepensesOfficiel (arrondis a l\'entier). Si aucun recapitulatif officiel n\'est visible sur le releve, mets exactement les memes valeurs que totalRecettes et totalDepenses.' + chr(10) +
        '6. Pour CHAQUE categorie de recettes et de depenses, liste dans "transactions" jusqu\'a 5 transactions individuelles les plus importantes qui la composent (libelle, montant, date)' + chr(10) +
        '7. Pour "compte", identifie le nom de la banque et/ou du compte TEL QU\'IL APPARAIT sur le releve (logo, en-tete, intitule de compte). Si tu ne trouves rien de clair, mets "' + nom_banque + '"' + chr(10) +
        '8. Identifie dans "prelevementsRecurrents" les charges probablement recurrentes/fixes de ce releve : abonnements, assurances, loyer, mensualites de credit, telephonie, energie, etc. (generalement des PRLV SEPA ou virements automatiques a montant fixe). Max 10, avec libelle et montant.' + chr(10) +
        chr(10) + 'Releve:' + chr(10) + _verifier_taille_et_tronquer(text)
    )
    msg = client.messages.create(
        model='claude-sonnet-4-6',
        max_tokens=4000,
        system='Tu es un expert-comptable. Tu reponds UNIQUEMENT avec du JSON valide, sans aucun texte avant ou apres, sans markdown, sans phrase d\'introduction ni de raisonnement visible (meme sur des releves complexes ou volumineux, va directement au JSON final). IMPORTANT: toutes les valeurs textuelles du JSON (labels de categories, commentaire, score_detail, titres et details des actions) doivent etre redigees dans la langue specifiee dans le prompt utilisateur.',
        messages=[{'role': 'user', 'content': prompt}]
    )
    raw = msg.content[0].text.replace('```json', '').replace('```', '').strip() if msg.content else ''
    stop_reason = getattr(msg, 'stop_reason', None)
    print('CLAUDE RESPONSE:', raw[:200], '| stop_reason:', stop_reason, '| nb_blocks:', len(msg.content), '| longueur_texte_source:', len(text))
    if not raw:
        raise ValueError('Reponse vide de Claude (stop_reason=' + str(stop_reason) + ', longueur texte source=' + str(len(text)) + ' caracteres)')
    try:
        return json.loads(raw)
    except json.JSONDecodeError:
        # L'IA a parfois ajoute une phrase de raisonnement avant/apres le JSON malgre la consigne.
        # On tente de recuperer uniquement le bloc JSON (premiere { a derniere }).
        start = raw.find('{')
        end = raw.rfind('}')
        if start != -1 and end != -1 and end > start:
            try:
                return json.loads(raw[start:end + 1])
            except json.JSONDecodeError:
                pass
        raise


def get_conseil_global(client, comptes, total_r, total_d, periode, langue='francais', patrimoine=None, devise='EUR'):
    comptes_str = chr(10).join(['- ' + c['nom'] + ' (' + c.get('periode', '') + '): recettes ' + str(c['totalRecettes']) + devise + ', depenses ' + str(c['totalDepenses']) + devise for c in comptes])
    net = total_r - total_d
    taux = round(net / total_r * 100) if total_r else 0
    langue_map = {'francais': 'French', 'english': 'English', 'espanol': 'Spanish', 'deutsch': 'German', 'italiano': 'Italian', 'portugues': 'Portuguese', 'chinese': 'Chinese', 'arabic': 'Arabic'}
    langue_name = langue_map.get(langue, 'French')

    patrimoine_str = ''
    if patrimoine:
        patrimoine_str = (
            chr(10) + 'ADDITIONAL CONTEXT - USER-DECLARED NET WORTH (beyond the bank flows above):' + chr(10) +
            '- Owned assets: ' + str(round(patrimoine['totalActifs'])) + devise + chr(10) +
            '- Savings/investments: ' + str(round(patrimoine['totalEpargnes'])) + devise + chr(10) +
            '- Remaining debts/loans: ' + str(round(patrimoine['totalDettes'])) + devise + chr(10) +
            '- Estimated net worth: ' + str(round(patrimoine['patrimoineNet'])) + devise + chr(10) +
            'Factor this into your score and advice: a tight monthly cash flow matters less if net worth is solid, and vice versa. Mention net worth explicitly if it materially changes the picture.' + chr(10)
        )

    prompt = (
        'You are a senior financial expert. Write ALL text EXCLUSIVELY in ' + langue_name + '. The currency of ALL amounts is ' + devise + ' - use this currency code (' + devise + ') everywhere you mention a monetary amount, NEVER write EUR or any other currency code.' + chr(10) +
        'Financial data for ' + periode + ':' + chr(10) +
        comptes_str + chr(10) +
        'TOTAL: income=' + str(total_r) + devise + ' expenses=' + str(total_d) + devise + ' net=' + str(net) + devise + ' savings_rate=' + str(taux) + '%' + chr(10) +
        patrimoine_str +
        'SCORING (be strict):' + chr(10) +
        '- 9-10: savings>30% AND positive net AND diversified income' + chr(10) +
        '- 7-8: savings 10-30% AND positive net' + chr(10) +
        '- 5-6: savings 0-10% OR slightly negative' + chr(10) +
        '- 3-4: savings -30% to 0%' + chr(10) +
        '- 1-2: savings < -30% OR deficit > 20% of income' + chr(10) +
        'Current savings rate=' + str(taux) + '% -> apply strictly.' + chr(10) +
        'PHRASE_CHOC: One single punchy sentence (max 15 words) that hits hard, using the ' + devise + ' currency code. Examples:' + chr(10) +
        '- "At this rate, your savings will be gone in 3 months."' + chr(10) +
        '- "You save the equivalent of a new iPhone every month."' + chr(10) +
        '- "Your biggest hidden expense: 4 forgotten subscriptions."' + chr(10) +
        '- "Warning: 2 of your 3 accounts are in the red."' + chr(10) +
        'Make it personal, specific with real numbers from the data, and impactful.' + chr(10) + chr(10) +
        'Return ONLY valid JSON, ALL text in ' + langue_name + ':' + chr(10) +
        '{"score":0,"score_detail":"sentence","phrase_choc":"impactful sentence with real numbers",' +
        '"actions":[{"priorite":1,"titre":"title","detail":"detail with numbers"},' +
        '{"priorite":2,"titre":"title","detail":"detail with numbers"},' +
        '{"priorite":3,"titre":"title","detail":"detail with numbers"}],' +
        '"commentaire":"2 sentences"}'
    )
    msg = client.messages.create(
        model='claude-sonnet-4-6',
        max_tokens=1000,
        messages=[{'role': 'user', 'content': prompt}]
    )
    raw = msg.content[0].text.replace('```json', '').replace('```', '').strip()
    return json.loads(raw)


@app.route('/create-checkout-session', methods=['POST'])
def create_checkout_session():
    try:
        data = request.get_json(force=True)
        email = (data.get('email') or '').strip().lower()
        if not email or '@' not in email:
            return jsonify({'error': 'Email invalide'}), 400

        conn = get_db()
        cur = conn.cursor()
        cur.execute('SELECT stripe_customer_id FROM users WHERE email = %s', (email,))
        row = cur.fetchone()
        cur.close()
        conn.close()

        customer_id = row['stripe_customer_id'] if row else None
        customer_valide = False
        if customer_id:
            try:
                stripe.Customer.retrieve(customer_id)
                customer_valide = True
            except stripe.error.InvalidRequestError:
                customer_valide = False

        if not customer_valide:
            customer = stripe.Customer.create(email=email)
            customer_id = customer.id
            upsert_user(email, stripe_customer_id=customer_id)
        else:
            # Deja un abonnement actif a ce prix ? On evite d'en creer un deuxieme (double facturation)
            abonnements_existants = stripe.Subscription.list(customer=customer_id, status='active', limit=10)
            for abo in abonnements_existants.data:
                for item in abo['items']['data']:
                    if item['price']['id'] == STRIPE_PRICE_ID:
                        return jsonify({'error': 'Vous etes deja abonne a BankAnalyzer Pro.', 'alreadySubscribed': True}), 409

        session = stripe.checkout.Session.create(
            customer=customer_id,
            mode='subscription',
            line_items=[{'price': STRIPE_PRICE_ID, 'quantity': 1}],
            subscription_data={'trial_period_days': 30},
            success_url=FRONTEND_URL + '?checkout=success',
            cancel_url=FRONTEND_URL + '?checkout=cancel',
        )
        return jsonify({'url': session.url})
    except Exception as e:
        print('ERREUR create_checkout_session:', str(e))
        return jsonify({'error': str(e)}), 500


def valeur_stripe(obj, cle, defaut=None):
    """Accede a un champ d'un objet Stripe (pas un dict standard, .get() n'est pas supporte)."""
    try:
        v = obj[cle]
        return v if v is not None else defaut
    except (KeyError, TypeError):
        return defaut


@app.route('/stripe-webhook', methods=['POST'])
def stripe_webhook():
    payload = request.get_data()
    sig_header = request.headers.get('Stripe-Signature')
    try:
        event = stripe.Webhook.construct_event(payload, sig_header, STRIPE_WEBHOOK_SECRET)
    except Exception as e:
        print('ERREUR signature webhook:', str(e))
        return jsonify({'error': 'Signature invalide'}), 400

    type_evenement = event['type']
    obj = event['data']['object']

    if type_evenement == 'checkout.session.completed':
        customer_id = valeur_stripe(obj, 'customer')
        subscription_id = valeur_stripe(obj, 'subscription')
        customer_details = valeur_stripe(obj, 'customer_details') or {}
        email = valeur_stripe(customer_details, 'email') or valeur_stripe(obj, 'customer_email')
        if email:
            upsert_user(email, stripe_customer_id=customer_id, stripe_subscription_id=subscription_id, subscription_status='active')
            print('Abonnement active pour', email)
        else:
            print('ATTENTION : aucun email trouve dans le checkout.session.completed')

    elif type_evenement in ('customer.subscription.updated', 'customer.subscription.deleted'):
        customer_id = valeur_stripe(obj, 'customer')
        statut = valeur_stripe(obj, 'status')
        nouveau_statut = 'active' if statut in ('active', 'trialing') else 'inactive'
        try:
            conn = get_db()
            cur = conn.cursor()
            cur.execute('SELECT email FROM users WHERE stripe_customer_id = %s', (customer_id,))
            row = cur.fetchone()
            cur.close()
            conn.close()
            if row:
                upsert_user(row['email'], subscription_status=nouveau_statut)
        except Exception as e:
            print('ERREUR maj abonnement webhook:', str(e))

    return jsonify({'received': True})


@app.route('/check-pro-status', methods=['POST'])
def check_pro_status():
    data = request.get_json(force=True)
    email = (data.get('email') or '').strip().lower()
    return jsonify({'isPro': est_pro(email)})


@app.route('/create-portal-session', methods=['POST'])
def create_portal_session():
    try:
        data = request.get_json(force=True)
        email = (data.get('email') or '').strip().lower()
        if not email or '@' not in email:
            return jsonify({'error': 'Email invalide'}), 400

        conn = get_db()
        cur = conn.cursor()
        cur.execute('SELECT stripe_customer_id FROM users WHERE email = %s', (email,))
        row = cur.fetchone()
        cur.close()
        conn.close()

        if not row or not row['stripe_customer_id']:
            return jsonify({'error': "Aucun abonnement trouve pour cet email"}), 404

        try:
            session = stripe.billing_portal.Session.create(
                customer=row['stripe_customer_id'],
                return_url=FRONTEND_URL,
            )
            return jsonify({'url': session.url})
        except stripe.error.InvalidRequestError as e:
            if 'No such customer' not in str(e):
                raise
            conn = get_db()
            cur = conn.cursor()
            cur.execute(
                "UPDATE users SET stripe_customer_id = NULL, stripe_subscription_id = NULL, subscription_status = 'inactive' WHERE email = %s",
                (email,)
            )
            conn.commit()
            cur.close()
            conn.close()
            return jsonify({'error': "Aucun abonnement actif trouve pour cet email. Vous pouvez souscrire a nouveau."}), 404
    except Exception as e:
        print('ERREUR create_portal_session:', str(e))
        return jsonify({'error': str(e)}), 500


@app.route('/admin/export-users', methods=['GET'])
def export_users():
    cle_fournie = request.args.get('key', '')
    cle_attendue = os.environ.get('ADMIN_SECRET', '')
    if not cle_attendue or cle_fournie != cle_attendue:
        return jsonify({'error': 'Non autorise'}), 403
    if not DATABASE_URL:
        return jsonify({'error': 'Base de donnees non configuree'}), 500
    try:
        import csv as _csv
        conn = get_db()
        cur = conn.cursor()
        cur.execute('''
            SELECT email, subscription_status, nb_analyses_mois_courant,
                   created_at, updated_at
            FROM users
            ORDER BY created_at DESC
        ''')
        lignes = cur.fetchall()
        cur.close()
        conn.close()

        tampon = io.StringIO()
        ecrivain = _csv.writer(tampon)
        ecrivain.writerow(['email', 'statut_abonnement', 'analyses_ce_mois', 'cree_le', 'maj_le'])
        for ligne in lignes:
            ecrivain.writerow([
                ligne['email'],
                ligne['subscription_status'],
                ligne['nb_analyses_mois_courant'],
                ligne['created_at'],
                ligne['updated_at'],
            ])

        return Response(
            tampon.getvalue(),
            mimetype='text/csv',
            headers={'Content-Disposition': 'attachment; filename=bankanalyzer-utilisateurs.csv'}
        )
    except Exception as e:
        print('ERREUR export_users:', str(e))
        return jsonify({'error': str(e)}), 500


@app.route('/admin/clean-stale-customers', methods=['GET'])
def clean_stale_customers():
    cle_fournie = request.args.get('key', '')
    cle_attendue = os.environ.get('ADMIN_SECRET', '')
    if not cle_attendue or cle_fournie != cle_attendue:
        return jsonify({'error': 'Non autorise'}), 403
    if not DATABASE_URL:
        return jsonify({'error': 'Base de donnees non configuree'}), 500
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("SELECT email, stripe_customer_id FROM users WHERE stripe_customer_id IS NOT NULL")
        lignes = cur.fetchall()
        cur.close()
        conn.close()

        emails_nettoyes = []
        emails_valides = []
        erreurs = []

        for ligne in lignes:
            email = ligne['email']
            customer_id = ligne['stripe_customer_id']
            try:
                stripe.Customer.retrieve(customer_id)
                emails_valides.append(email)
            except stripe.error.InvalidRequestError as e:
                if 'No such customer' in str(e):
                    conn2 = get_db()
                    cur2 = conn2.cursor()
                    cur2.execute(
                        "UPDATE users SET stripe_customer_id = NULL, stripe_subscription_id = NULL, subscription_status = 'inactive' WHERE email = %s",
                        (email,)
                    )
                    conn2.commit()
                    cur2.close()
                    conn2.close()
                    emails_nettoyes.append(email)
                else:
                    erreurs.append({'email': email, 'erreur': str(e)})

        return jsonify({
            'total_verifies': len(lignes),
            'valides': len(emails_valides),
            'nettoyes': len(emails_nettoyes),
            'emails_nettoyes': emails_nettoyes,
            'erreurs': erreurs,
        })
    except Exception as e:
        print('ERREUR clean_stale_customers:', str(e))
        return jsonify({'error': str(e)}), 500


@app.route('/analyze', methods=['POST'])
def analyze():
    try:
        return _analyze_impl()
    except Exception as e:
        import traceback
        print('ERREUR FATALE /analyze:', str(e))
        print(traceback.format_exc())
        return jsonify({'error': "Erreur serveur : " + str(e)[:300]}), 500


def _analyze_impl():
    langue = request.form.get('langue', 'français')
    files = request.files.getlist('files')

    email = (request.form.get('email') or '').strip().lower()
    if not email or '@' not in email:
        return jsonify({'error': 'Adresse email requise pour lancer une analyse.'}), 400
    upsert_user(email)

    mode_devise = request.form.get('modeDevise', 'unique')
    devise_unique = (request.form.get('deviseUnique') or 'EUR').upper().strip()
    devise_reference = (request.form.get('deviseReference') or 'EUR').upper().strip()
    devise_principale = devise_unique if mode_devise == 'unique' else devise_reference

    devises_raw = request.form.get('devises', '')
    taux_par_devise = {}
    if devises_raw:
        try:
            liste_taux = json.loads(devises_raw)
            for item in liste_taux:
                code = (item.get('code') or '').upper().strip()
                taux = to_num(item.get('taux'))
                if code and taux > 0:
                    taux_par_devise[code] = taux
        except (json.JSONDecodeError, AttributeError, TypeError):
            taux_par_devise = {}

    def convertir_montant_patrimoine(montant, devise_entree, taux_ligne):
        m = to_num(montant)
        devise_entree = (devise_entree or devise_principale).upper().strip()
        if devise_entree == devise_principale:
            return m
        t = to_num(taux_ligne)
        if t > 0:
            return m * t
        if devise_entree in taux_par_devise:
            return m * taux_par_devise[devise_entree]
        return m  # aucun taux disponible : on garde la valeur brute (imprecis mais on ne bloque pas)

    patrimoine_raw = request.form.get('patrimoine', '')
    patrimoine_resume = None
    if patrimoine_raw:
        try:
            patrimoine = json.loads(patrimoine_raw)
            emprunts = patrimoine.get('emprunts', []) or []
            epargnes = patrimoine.get('epargnes', []) or []
            actifs = patrimoine.get('actifs', []) or []
            if emprunts or epargnes or actifs:
                total_actifs = sum(convertir_montant_patrimoine(a.get('valeur'), a.get('devise'), a.get('tauxConversion')) for a in actifs)
                total_epargnes = sum(convertir_montant_patrimoine(e.get('montant'), e.get('devise'), e.get('tauxConversion')) for e in epargnes)
                total_dettes = sum(convertir_montant_patrimoine(e.get('capitalRestantDu') or e.get('montantInitial'), e.get('devise'), e.get('tauxConversion')) for e in emprunts)
                patrimoine_resume = {
                    'totalActifs': total_actifs,
                    'totalEpargnes': total_epargnes,
                    'totalDettes': total_dettes,
                    'patrimoineNet': total_actifs + total_epargnes - total_dettes,
                    'emprunts': emprunts,
                    'epargnes': epargnes,
                    'actifs': actifs
                }
        except (json.JSONDecodeError, AttributeError, TypeError):
            patrimoine_resume = None

    if not files:
        f = request.files.get('file')
        if f:
            files = [f]
        else:
            return jsonify({'error': 'Aucun fichier recu'}), 400
    files = files[:60]

    utilisateur_pro = est_pro(email)

    client = anthropic.Anthropic()

    # Etape 1 : extraction et decoupage par mois de chaque fichier (rapide, local, sequentiel)
    fichiers_prepares = []
    fichiers_ignores = []
    for file in files:
        filename = file.filename.lower()
        file_bytes = file.read()
        nom_banque = file.filename.replace('.pdf', '').replace('.csv', '').replace('.xlsx', '')[:30]
        try:
            if filename.endswith('.pdf'):
                texte_pdf_extrait = extract_text_from_pdf(file_bytes)
                blocs = [{'periode': None, 'text': texte_pdf_extrait, 'totauxVerifies': calculer_totaux_verifies_pdf(texte_pdf_extrait)}]
            elif filename.endswith('.csv'):
                blocs = decouper_par_mois_csv(file_bytes)
            elif filename.endswith(('.xlsx', '.xls')):
                blocs = decouper_par_mois_excel(file_bytes)
            else:
                fichiers_ignores.append({'nom': file.filename, 'raison': 'Format non supporte'})
                continue
        except Exception as e:
            print('ERREUR extraction', file.filename, str(e))
            fichiers_ignores.append({'nom': file.filename, 'raison': "Fichier illisible/corrompu : " + str(e)[:150]})
            continue
        if not blocs or all(not b['text'].strip() for b in blocs):
            fichiers_ignores.append({'nom': file.filename, 'raison': 'Aucun texte extrait (PDF scanne/image ?)'})
            continue
        for bloc in blocs:
            texte_bloc = bloc['text']
            if not texte_bloc.strip():
                continue
            periode = bloc['periode'] or get_periode(texte_bloc)
            nom_bloc = nom_banque if len(blocs) == 1 else (nom_banque + ' - ' + periode if periode else nom_banque)
            fichiers_prepares.append({'nom': nom_bloc, 'nomFichier': file.filename, 'periode': periode, 'text': texte_bloc, 'totauxVerifies': bloc.get('totauxVerifies'), 'transactionsStructurees': bloc.get('transactionsStructurees')})

    if not fichiers_prepares:
        return jsonify({'error': 'Aucun releve analyse'}), 500

    demande_features_pro = len(fichiers_prepares) > 1 or bool(patrimoine_resume) or mode_devise == 'multiple'
    if demande_features_pro and not utilisateur_pro:
        return jsonify({
            'error': "Cette analyse utilise une fonctionnalite BankAnalyzer Pro (plusieurs fichiers, plusieurs mois detectes dans un meme fichier, patrimoine ou plusieurs devises). Passez a l'offre Pro pour y acceder.",
            'requiresPro': True
        }), 402

    if utilisateur_pro:
        quota_ok, nb_utilisees = verifier_et_incrementer_quota_pro(email)
        if not quota_ok:
            return jsonify({
                'error': "Vous avez atteint votre quota de %d analyses Pro pour ce mois-ci. Il sera reinitialise le mois prochain." % QUOTA_ANALYSES_PRO_PAR_MOIS,
                'quotaDepasse': True
            }), 402

    # Etape 2 : appels IA en parallele (par lots pour respecter les limites de debit de l'API)
    comptes = []
    TAILLE_LOT = 8

    avertissements = []
    SEUIL_ECART = 0.03  # 3% d'ecart tolere avant d'avertir

    def analyser_un_fichier(f):
        derniere_erreur = None
        for tentative in range(2):
            try:
                totaux_verifies = f.get('totauxVerifies')
                data = analyse_releve(client, f['text'], f['nom'], langue, totaux_verifies=totaux_verifies)
                nom_detecte = (data.get('compte') or '').strip()
                data['nom'] = nom_detecte if nom_detecte else f['nom']
                data['periode'] = f['periode']
                _appliquer_categorisation_exhaustive_si_necessaire(client, f, data, langue)

                if totaux_verifies:
                    # Filet de securite : on ecrase les totaux de l'IA par les
                    # totaux calcules en Python, au cas ou l'instruction du
                    # prompt n'aurait pas ete suivie a la lettre.
                    data['totalRecettes'] = totaux_verifies['totalRecettes']
                    data['totalDepenses'] = totaux_verifies['totalDepenses']
                    data['totalRecettesOfficiel'] = totaux_verifies['totalRecettes']
                    data['totalDepensesOfficiel'] = totaux_verifies['totalDepenses']
                    if 'soldeDepart' in totaux_verifies:
                        # Certains formats (Revolut) fournissent un vrai solde
                        # d'ouverture/cloture fiable : on l'utilise tel quel.
                        data['soldeDepart'] = totaux_verifies['soldeDepart']
                        data['soldeArrivee'] = totaux_verifies['soldeArrivee']
                    else:
                        # Sinon (CSV/XLSX avec colonnes Debit/Credit uniquement),
                        # on ne sait pas soldeDepart/soldeArrivee : on le dit
                        # explicitement (null) plutot que de laisser 0.
                        data['soldeDepart'] = None
                        data['soldeArrivee'] = None

                data = _completer_avec_categorie_autres(data)

                tr = to_num(data.get('totalRecettes', 0))
                td = to_num(data.get('totalDepenses', 0))
                tr_off = to_num(data.get('totalRecettesOfficiel', tr))
                td_off = to_num(data.get('totalDepensesOfficiel', td))
                ecarts = []
                if tr_off > 0 and abs(tr - tr_off) / tr_off > SEUIL_ECART:
                    ecarts.append('recettes calculees=' + str(round(tr)) + ' vs officiel=' + str(round(tr_off)))
                if td_off > 0 and abs(td - td_off) / td_off > SEUIL_ECART:
                    ecarts.append('depenses calculees=' + str(round(td)) + ' vs officiel=' + str(round(td_off)))
                if ecarts:
                    avertissements.append({
                        'nom': data['nom'] + ' (' + f['periode'] + ')',
                        'raison': 'Ecart avec le recapitulatif officiel du releve : ' + ' ; '.join(ecarts)
                    })

                _logger_banque_detectee(data.get('nom'), f.get('nomFichier'), f.get('periode'))

                return data
            except Exception as e:
                derniere_erreur = e
        raise derniere_erreur

    for i in range(0, len(fichiers_prepares), TAILLE_LOT):
        lot = fichiers_prepares[i:i + TAILLE_LOT]
        with ThreadPoolExecutor(max_workers=TAILLE_LOT) as executor:
            futures = {executor.submit(analyser_un_fichier, f): f for f in lot}
            for future in as_completed(futures):
                f = futures[future]
                try:
                    comptes.append(future.result())
                except Exception as e:
                    print('ERREUR releve', f['nom'], str(e))
                    err_msg = str(e)[:150]
                    fichiers_ignores.append({'nom': f['nomFichier'], 'raison': "Echec de l'analyse IA : " + err_msg})
                    continue

    if not comptes:
        return jsonify({'error': 'Aucun releve analyse', 'fichiersIgnores': fichiers_ignores}), 500

    # --- Devise : mode choisi explicitement par l'utilisateur ---
    def convertir_compte_devise(compte, taux, nouvelle_devise):
        c = dict(compte)
        for champ in ['totalRecettes', 'totalDepenses', 'soldeDepart', 'soldeArrivee']:
            if champ in c:
                c[champ] = to_num(c[champ]) * taux
        for cle in ['recettes', 'depenses']:
            nouvelles = []
            for item in c.get(cle, []):
                item2 = dict(item)
                item2['montant'] = to_num(item2.get('montant', 0)) * taux
                item2['transactions'] = [
                    dict(t, montant=to_num(t.get('montant', 0)) * taux) for t in item.get('transactions', [])
                ]
                nouvelles.append(item2)
            c[cle] = nouvelles
        c['top5depenses'] = [
            dict(t, montant=to_num(t.get('montant', 0)) * taux) for t in c.get('top5depenses', [])
        ]
        c['devise'] = nouvelle_devise
        return c

    if mode_devise == 'unique':
        # L'utilisateur affirme que tout est deja dans une seule et meme devise : on force cette devise partout
        for c in comptes:
            c['devise'] = devise_unique
        devise_principale = devise_unique
    else:
        devise_principale = devise_reference
        comptes_convertis = []
        comptes_hors_devise = []
        for c in comptes:
            dv = (c.get('devise') or 'EUR').upper().strip()
            if dv == devise_principale:
                comptes_convertis.append(c)
            elif dv in taux_par_devise:
                comptes_convertis.append(convertir_compte_devise(c, taux_par_devise[dv], devise_principale))
            else:
                comptes_hors_devise.append(c)
        for c in comptes_hors_devise:
            fichiers_ignores.append({
                'nom': c.get('nom', '?') + ' (' + c.get('periode', '') + ')',
                'raison': 'Devise ' + (c.get('devise') or '?') + ' - aucun taux de conversion fourni vers ' + devise_principale
            })
        comptes = comptes_convertis

    if not comptes:
        return jsonify({'error': 'Aucun compte ne correspond a la devise choisie, aucune analyse coherente possible', 'fichiersIgnores': fichiers_ignores}), 500

    # --- Detection automatique multi-mois vs multi-comptes ---
    periodes_uniques = sorted(
        set(c.get('periode', 'Periode inconnue') for c in comptes),
        key=periode_sort_key
    )
    is_multi_mois = len(periodes_uniques) > 1

    def _sommer_solde(liste_comptes, champ):
        """Comme sum(to_num(...)) mais renvoie None si AUCUN compte de la
        liste n'a de valeur connue pour ce champ (evite qu'un solde
        veritablement inconnu ne se transforme silencieusement en 0)."""
        valeurs = [c.get(champ) for c in liste_comptes]
        if valeurs and all(v is None for v in valeurs):
            return None
        return sum(to_num(v) for v in valeurs)

    evolution = []
    if is_multi_mois:
        premiere_periode = periodes_uniques[0]
        derniere_periode = periodes_uniques[-1]
        comptes_premiere = [c for c in comptes if c.get('periode', 'Periode inconnue') == premiere_periode]
        comptes_derniere = [c for c in comptes if c.get('periode', 'Periode inconnue') == derniere_periode]
        solde_depart = _sommer_solde(comptes_premiere, 'soldeDepart')
        solde_arrivee = _sommer_solde(comptes_derniere, 'soldeArrivee')
        periode_label = premiere_periode + ' -> ' + derniere_periode

        if solde_depart is not None:
            # Point de depart : solde d'ouverture du tout premier mois, avant tout mouvement
            evolution.append({
                'periode': 'Debut ' + premiere_periode,
                'totalRecettes': 0,
                'totalDepenses': 0,
                'net': 0,
                'soldeArrivee': solde_depart
            })
            for p in periodes_uniques:
                comptes_p = [c for c in comptes if c.get('periode', 'Periode inconnue') == p]
                tr_p = sum(to_num(c.get('totalRecettes', 0)) for c in comptes_p)
                td_p = sum(to_num(c.get('totalDepenses', 0)) for c in comptes_p)
                sa_p = _sommer_solde(comptes_p, 'soldeArrivee')
                evolution.append({
                    'periode': p,
                    'totalRecettes': tr_p,
                    'totalDepenses': td_p,
                    'net': tr_p - td_p,
                    'soldeArrivee': sa_p if sa_p is not None else 0
                })
        # Si solde_depart est None (aucune donnee de solde disponible pour ce
        # type d'export), evolution reste vide : le frontend affiche alors son
        # message "en attente de plus de mois" existant plutot qu'un
        # graphique plat trompeur -- aucune modification frontend requise.
    else:
        solde_depart = _sommer_solde(comptes, 'soldeDepart')
        solde_arrivee = _sommer_solde(comptes, 'soldeArrivee')
        periode_label = periodes_uniques[0] if periodes_uniques else 'Periode inconnue'

    # Vue d'ensemble = toujours la somme de TOUS les mois/comptes envoyes
    total_r = sum(to_num(c.get('totalRecettes', 0)) for c in comptes)
    total_d = sum(to_num(c.get('totalDepenses', 0)) for c in comptes)

    # Table de synonymes connus : mots de racines differentes designant la
    # meme realite (ex. "echeance" et "remboursement" pour un pret).
    # Cle et valeur en minuscules, sans accents. Extensible au fil du temps.
    _SYNONYMES_CATEGORIE = {
        'echeance': 'remboursement',
    }
    _MOTS_OUTILS_CATEGORIE = {'de', 'du', 'des', 'la', 'le', 'les', 'd', 'l', 'et', 'un', 'une'}

    def _normaliser_cle_categorie(label):
        # Cle de regroupement insensible aux accents, a la ponctuation, aux
        # mots-outils et a quelques synonymes connus. Deux libelles qui
        # donnent la meme cle sont consideres comme la meme categorie.
        s = sans_accents(str(label or '')).lower()
        s = re.sub(r"[^a-z0-9\s]", ' ', s)
        mots = [m for m in s.split() if m and m not in _MOTS_OUTILS_CATEGORIE]
        mots = [_SYNONYMES_CATEGORIE.get(m, m) for m in mots]
        return ' '.join(sorted(mots))

    _appliquer_harmonisation_categories_si_necessaire(client, comptes, langue)

    all_rec = {}
    all_rec_tx = {}
    all_rec_labels = {}
    all_dep = {}
    all_dep_tx = {}
    all_dep_labels = {}
    for c in comptes:
        for r in c.get('recettes', []):
            label_r = r.get('label')
            cle = _normaliser_cle_categorie(label_r)
            all_rec_labels.setdefault(cle, label_r or 'Autres')
            all_rec[cle] = all_rec.get(cle, 0) + to_num(r.get('montant', 0))
            all_rec_tx.setdefault(cle, []).extend(r.get('transactions', []))
        for d in c.get('depenses', []):
            label_d = d.get('label')
            cle = _normaliser_cle_categorie(label_d)
            all_dep_labels.setdefault(cle, label_d or 'Autres')
            all_dep[cle] = all_dep.get(cle, 0) + to_num(d.get('montant', 0))
            all_dep_tx.setdefault(cle, []).extend(d.get('transactions', []))

    # On reconvertit les cles normalisees (illisibles) vers le premier
    # libelle original rencontre, pour l'affichage final.
    all_rec = {all_rec_labels[k]: v for k, v in all_rec.items()}
    all_rec_tx = {all_rec_labels[k]: v for k, v in all_rec_tx.items()}
    all_dep = {all_dep_labels[k]: v for k, v in all_dep.items()}
    all_dep_tx = {all_dep_labels[k]: v for k, v in all_dep_tx.items()}

    def normaliser_montants(transactions):
        out = []
        for t in transactions:
            t2 = dict(t)
            t2['montant'] = to_num(t2.get('montant', 0))
            out.append(t2)
        return out

    def top_transactions(liste):
        return normaliser_montants(sorted(liste, key=lambda x: -to_num(x.get('montant', 0)))[:8])

    def _finaliser_categories_fusionnees(items_dict, items_tx, total, limite):
        """Construit la liste finale de categories fusionnees (max
        `limite`), TOUJOURS reconciliee exactement avec `total`. Une place
        est toujours reservee pour 'Autres' si necessaire, plutot que de
        risquer qu'elle soit ecartee par un simple tri par montant."""
        def _est_autres(label):
            return sans_accents(str(label or '').strip().lower()) in ('autres', 'autre', 'divers')

        montant_autres_deja = sum(v for k, v in items_dict.items() if _est_autres(k))
        transactions_autres_deja = []
        for k in items_dict:
            if _est_autres(k):
                transactions_autres_deja.extend(items_tx.get(k, []))
        items = sorted(
            [(k, v) for k, v in items_dict.items() if not _est_autres(k)],
            key=lambda kv: -kv[1]
        )

        principales = items[:limite]
        resultat = [
            {'label': k, 'montant': round(v, 2), 'transactions': normaliser_montants(items_tx.get(k, []))}
            for k, v in principales
        ]

        somme_affichee = sum(x['montant'] for x in resultat)
        ecart = round(total - somme_affichee, 2)
        if ecart > 1 or montant_autres_deja > 1:
            ecart_total = round(ecart, 2) if ecart > 1 else round(montant_autres_deja, 2)
            transactions_autres_finales = normaliser_montants(transactions_autres_deja)
            if len(resultat) < limite:
                resultat.append({'label': 'Autres', 'montant': ecart_total, 'transactions': transactions_autres_finales})
            else:
                idx_min = min(range(len(resultat)), key=lambda i: resultat[i]['montant'])
                transactions_fusionnees = resultat[idx_min].get('transactions', []) + transactions_autres_finales
                resultat[idx_min] = {
                    'label': 'Autres',
                    'montant': round(resultat[idx_min]['montant'] + ecart_total, 2),
                    'transactions': transactions_fusionnees,
                }
        return resultat

    rec_global = _finaliser_categories_fusionnees(all_rec, all_rec_tx, total_r, 5)
    dep_global = _finaliser_categories_fusionnees(all_dep, all_dep_tx, total_d, 7)

    try:
        conseil = get_conseil_global(client, comptes, total_r, total_d, periode_label, langue, patrimoine_resume, devise_principale)
    except Exception as e:
        print('AVERTISSEMENT get_conseil_global a echoue:', repr(e))
        conseil = {'score': 5, 'score_detail': 'Analyse partielle', 'actions': [], 'commentaire': 'Analyse disponible.'}

    all_top5 = []
    for c in comptes:
        all_top5.extend(c.get('top5depenses', []))
    all_top5 = sorted(all_top5, key=lambda x: -to_num(x.get('montant', 0)))[:5]
    all_top5 = normaliser_montants(all_top5)

    charges_dict = {}
    for c in comptes:
        for p in (c.get('prelevementsRecurrents') or []):
            label = (p.get('libelle') or '').strip()
            if not label:
                continue
            montant = to_num(p.get('montant', 0))
            if montant <= 0:
                continue
            key = _normaliser_cle_categorie(label)
            if key not in charges_dict:
                charges_dict[key] = {'libelle': label, 'total': 0.0, 'mois': set(), 'montants': []}
            charges_dict[key]['total'] += montant
            charges_dict[key]['mois'].add(c.get('periode', ''))
            charges_dict[key]['montants'].append(montant)

    charges_fixes_liste = []
    for v in charges_dict.values():
        nb_mois = len(v['mois'])
        montant_moyen = v['total'] / len(v['montants']) if v['montants'] else 0
        charges_fixes_liste.append({
            'libelle': v['libelle'],
            'montantMoyen': montant_moyen,
            'montantTotal': v['total'],
            'nbMois': nb_mois
        })
    charges_fixes_liste = sorted(charges_fixes_liste, key=lambda x: -x['montantTotal'])[:15]
    total_charges_fixes = sum(x['montantTotal'] for x in charges_fixes_liste)
    pourcentage_charges_fixes = round(total_charges_fixes / total_d * 100) if total_d else 0
    charges_fixes_resume = {
        'liste': charges_fixes_liste,
        'total': total_charges_fixes,
        'pourcentage': pourcentage_charges_fixes
    }

    result = {
        'periode': periode_label,
        'devise': devise_principale,
        'isMultiMois': is_multi_mois,
        'evolution': evolution,
        'fichiersIgnores': fichiers_ignores,
        'avertissements': avertissements,
        'chargesFixes': charges_fixes_resume,
        'patrimoine': patrimoine_resume,
        'totalRecettes': total_r,
        'totalDepenses': total_d,
        'soldeDepart': solde_depart,
        'soldeArrivee': solde_arrivee,
        'top5depenses': all_top5,
        'recettes': rec_global,
        'depenses': dep_global,
        'phrase_choc': conseil.get('phrase_choc', ''),
        'score': conseil.get('score', 5),
        'score_detail': conseil.get('score_detail', ''),
        'actions': conseil.get('actions', []),
        'commentaire': conseil.get('commentaire', ''),
        'comptes': comptes
    }

    if not utilisateur_pro:
        result['chargesFixes'] = None
        result['isPro'] = False
        for cle in ['recettes', 'depenses']:
            for item in result.get(cle, []):
                item['transactions'] = []
    else:
        result['isPro'] = True

    print('PHRASE_CHOC:', result.get('phrase_choc', 'VIDE'))
    return jsonify(result)


if __name__ == '__main__':
    app.run(host='0.0.0.0', port=int(os.environ.get('PORT', 5001)))
